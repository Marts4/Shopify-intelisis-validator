"""
Generador de reportes con datos crudos de Shopify e Intelisis.
Crea archivos Excel separados por fecha antes del procesamiento.
"""

import os
from pathlib import Path
from typing import List, Dict
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from config import settings
from src.core.models import Order
from src.core.exceptions import ReportGenerationError
from src.utils import get_logger


class RawDataReporter:
    """
    Genera reportes Excel con datos crudos de las consultas.
    Crea archivos separados por fecha: consulta_shop_YYYYMMDD.xlsx y consulta_intelisis_YYYYMMDD.xlsx
    """

    def __init__(self):
        self.logger = get_logger('RAW_DATA_REPORTER')

        # Directorio de salida
        self.output_dir = settings.excel_report.full_output_path.parent

        # Estilos
        self.header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
        self.header_font = Font(
            name='Arial',
            size=11,
            bold=True,
            color="FFFFFF"
        )
        self.header_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

        # Tracking de archivos generados
        self.generated_files = {
            'shopify': [],
            'intelisis': []
        }

    def save_shopify_raw_data(self, raw_data: List[Dict], date: str):
        """
        Guarda datos crudos de Shopify en Excel con TODOS los campos originales.
        Crea un archivo por fecha: consulta_shop_YYYYMMDD.xlsx

        Args:
            raw_data: Lista de diccionarios con los datos originales de la API
            date: Fecha consultada (formato YYYY-MM-DD)
        """
        if not raw_data:
            self.logger.warning("No hay datos de Shopify para guardar")
            return

        # Generar nombre de archivo con fecha
        date_suffix = date.replace('-', '')  # 2025-11-01 -> 20251101
        filename = f"consulta_shop_{date_suffix}.xlsx"
        filepath = self.output_dir / filename

        self.logger.info(f"Guardando datos crudos de Shopify: {len(raw_data)} órdenes")
        self.logger.info(f"Archivo: {filename}")

        try:
            # Convertir a DataFrame directamente de los datos crudos
            df = pd.DataFrame(raw_data)

            # Crear Excel con formato
            self._save_dataframe_with_format(
                df=df,
                path=filepath,
                sheet_name="Shopify_Raw"
            )

            self.generated_files['shopify'].append({
                'path': str(filepath),
                'date': date,
                'records': len(df)
            })

            self.logger.info(f"✓ Datos guardados: {filename} ({len(df)} órdenes)")

        except Exception as e:
            raise ReportGenerationError(f"Error guardando datos Shopify: {str(e)}")

    def save_intelisis_raw_data(self, raw_data: List[Dict], date: str):
        """
        Guarda datos crudos de Intelisis en Excel con TODOS los campos originales.
        Crea un archivo por fecha: consulta_intelisis_YYYYMMDD.xlsx

        Args:
            raw_data: Lista de diccionarios con los datos filtrados de Intelisis
            date: Fecha consultada (formato YYYY-MM-DD)
        """
        if not raw_data:
            self.logger.warning("No hay datos de Intelisis para guardar")
            return

        # Generar nombre de archivo con fecha
        date_suffix = date.replace('-', '')  # 2025-11-01 -> 20251101
        filename = f"consulta_intelisis_{date_suffix}.xlsx"
        filepath = self.output_dir / filename

        self.logger.info(f"Guardando datos crudos de Intelisis: {len(raw_data)} registros")
        self.logger.info(f"Archivo: {filename}")

        try:
            # Convertir a DataFrame directamente de los datos crudos
            df = pd.DataFrame(raw_data)

            # Reordenar columnas: campos clave primero
            priority_cols = [
                'Sucursal', 'Mov', 'Adicional5', 'NombreCliente',
                'Cantidad', 'vtcTotalNeto', 'Adicional2'
            ]

            # Columnas prioritarias que existen
            existing_priority = [col for col in priority_cols if col in df.columns]
            # Resto de columnas
            remaining_cols = [col for col in df.columns if col not in existing_priority]

            # Reordenar
            df = df[existing_priority + remaining_cols]

            # Crear Excel con formato
            self._save_dataframe_with_format(
                df=df,
                path=filepath,
                sheet_name="Intelisis_Raw"
            )

            self.generated_files['intelisis'].append({
                'path': str(filepath),
                'date': date,
                'records': len(df)
            })

            self.logger.info(f"✓ Datos guardados: {filename} ({len(df)} registros)")

        except Exception as e:
            raise ReportGenerationError(f"Error guardando datos Intelisis: {str(e)}")

    def _save_dataframe_with_format(self, df: pd.DataFrame, path: Path, sheet_name: str):
        """
        Guarda DataFrame en Excel con formato profesional.

        Args:
            df: DataFrame a guardar
            path: Ruta del archivo
            sheet_name: Nombre de la hoja
        """
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Escribir headers
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment

        # Escribir datos
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                # Convertir valores especiales a string legible
                try:
                    if pd.isna(value):
                        value = ""
                    elif isinstance(value, (list, dict)):
                        value = str(value)
                    elif hasattr(value, '__iter__') and not isinstance(value, str):
                        # Convertir cualquier iterable (excepto strings) a string
                        value = str(value)
                except (ValueError, TypeError):
                    # Si falla la conversión, usar string vacío
                    value = ""

                ws.cell(row=row_idx, column=col_idx, value=value)

        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Auto-filtro en headers
        ws.auto_filter.ref = ws.dimensions

        # Congelar primera fila
        ws.freeze_panes = 'A2'

        # Guardar
        wb.save(path)

    def get_summary(self) -> dict:
        """
        Obtiene resumen de archivos generados.

        Returns:
            Diccionario con información de los archivos
        """
        return {
            'shopify': self.generated_files['shopify'],
            'intelisis': self.generated_files['intelisis']
        }