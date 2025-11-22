"""
Generador de reportes en formato Excel.
Crea archivos con formato alternado y columnas específicas.
"""

import os
from pathlib import Path
from typing import List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from config import settings
from src.core.models import ValidationResult, ValidationSummary
from src.core.exceptions import ReportGenerationError
from .base_reporter import BaseReporter


class ExcelReporter(BaseReporter):
    """
    Genera reportes en formato Excel con formato alternado.
    Agrupa visualmente por plataforma y order_number.
    """
    
    def __init__(self):
        super().__init__('EXCEL_REPORTER')
        self.output_path = settings.excel_report.full_output_path
        
        # Estilos para el Excel
        self.gray_fill = PatternFill(
            start_color="A4A4A4",
            end_color="A4A4A4",
            fill_type="solid"
        )
        self.no_fill = PatternFill(fill_type=None)
    
    def _prepare(
        self, 
        results: List[ValidationResult], 
        summary: ValidationSummary,
        date_description: str
    ):
        """Crea el directorio de salida si no existe"""
        try:
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            self.logger.debug(f"Directorio de salida verificado: {self.output_path.parent}")
        except Exception as e:
            raise ReportGenerationError(f"Error creando directorio: {str(e)}")
    
    def _create_report(
        self, 
        results: List[ValidationResult], 
        summary: ValidationSummary,
        date_description: str
    ):
        """
        Genera el archivo Excel con los resultados.
        
        Args:
            results: Lista de resultados de validación
            summary: Resumen estadístico
            date_description: Descripción de fechas
        """
        if not results:
            self.logger.warning("No hay resultados para exportar a Excel")
            return
        
        # Convertir resultados a DataFrame
        df = self._results_to_dataframe(results)
        
        self.logger.info(f"Creando archivo Excel: {self.output_path}")
        self.logger.info(f"Total registros: {len(df)}")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Validacion"
        
        # Escribir headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Crear mapa de colores por combinación única de plataforma_order_number
        color_map = self._create_color_map(df)
        
        # Escribir datos con formato alternado
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            key = f"{row['plataforma']}_{row['order_number']}"
            fill_color = color_map.get(key, self.no_fill)
            
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill_color
        
        # Ajustar anchos de columna (opcional, mejora legibilidad)
        self._adjust_column_widths(ws, df)
        
        # Guardar archivo
        try:
            wb.save(self.output_path)
            self.logger.info(f"✓ Archivo guardado: {self.output_path}")
        except Exception as e:
            raise ReportGenerationError(f"Error guardando Excel: {str(e)}")
    
    def _results_to_dataframe(self, results: List[ValidationResult]) -> pd.DataFrame:
        """
        Convierte lista de ValidationResult a DataFrame.
        
        Args:
            results: Lista de resultados
            
        Returns:
            DataFrame con todas las columnas necesarias
        """
        data = [result.to_dict() for result in results]
        df = pd.DataFrame(data)
        
        # Asegurar orden de columnas consistente
        column_order = [
            'plataforma',
            'id',
            'order_number',
            'created_at',
            'financial_status',
            'customer_name',
            'nombre_intelisis',
            'quantity',
            'cantidad_intelisis',
            'registros',
            'registros_intelisis',
            'total_final',
            'total_intelisis',
            'coordinates',
            'coordenadas_intelisis',
            'observaciones'
        ]
        
        # Reordenar columnas (si existen)
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]
        
        return df
    
    def _create_color_map(self, df: pd.DataFrame) -> dict:
        """
        Crea un mapa de colores alternados por plataforma_order_number.
        
        Args:
            df: DataFrame con los datos
            
        Returns:
            Diccionario {key: fill_color}
        """
        # Obtener claves únicas preservando el orden
        df['_key'] = df['plataforma'] + '_' + df['order_number'].astype(str)
        unique_keys = df['_key'].unique()
        
        # Asignar colores alternados
        color_map = {}
        for idx, key in enumerate(unique_keys):
            color_map[key] = self.gray_fill if idx % 2 == 1 else self.no_fill
        
        return color_map
    
    def _adjust_column_widths(self, ws, df: pd.DataFrame):
        """
        Ajusta el ancho de las columnas basándose en el contenido.
        
        Args:
            ws: Worksheet de openpyxl
            df: DataFrame con los datos
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Establecer ancho (con un máximo de 50)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _finalize(
        self, 
        results: List[ValidationResult], 
        summary: ValidationSummary,
        date_description: str
    ):
        """Verifica que el archivo fue creado correctamente"""
        if not self.output_path.exists():
            raise ReportGenerationError(
                f"El archivo Excel no fue creado: {self.output_path}"
            )
        
        file_size = self.output_path.stat().st_size
        self.logger.info(f"Tamaño del archivo: {file_size:,} bytes")
    
    def get_output_path(self) -> Path:
        """
        Retorna la ruta del archivo generado.
        
        Returns:
            Path del archivo Excel
        """
        return self.output_path